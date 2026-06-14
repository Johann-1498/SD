"""
Catalog reader.
Reads product catalog from Excel (.xlsx) when openpyxl is available,
otherwise from a CSV fallback. Produces a list of dicts:
    [{'sku': ..., 'name': ..., 'price': ..., 'warehouse': ..., 'quantity': ...}]
"""
import csv
import logging
import os

logger = logging.getLogger(__name__)

OPENPYXL_AVAILABLE = True
try:
    from openpyxl import load_workbook
except ImportError:
    OPENPYXL_AVAILABLE = False


def read_catalog(path):
    if not os.path.exists(path):
        raise FileNotFoundError(f'Catalog file not found: {path}')
    if path.lower().endswith('.xlsx') and OPENPYXL_AVAILABLE:
        return _read_xlsx(path)
    csv_path = path
    if path.lower().endswith('.xlsx'):
        csv_path = path[:-5] + '.csv'
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f'Neither .xlsx nor fallback .csv found for {path}')
    return _read_csv(csv_path)


def _read_xlsx(path):
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h is not None else '' for h in rows[0]]
    records = []
    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue
        record = {header[i]: row[i] for i in range(len(header))}
        records.append(record)
    logger.info('Read %d records from xlsx', len(records))
    return records


def _read_csv(path):
    records = []
    with open(path, newline='', encoding='utf-8') as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            records.append(dict(row))
    logger.info('Read %d records from csv', len(records))
    return records
